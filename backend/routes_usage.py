import csv
import os
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status

from .time_utils import (
    china_today,
    now_china_as_utc,
    parse_naive_as_china_then_utc,
    utc_naive_to_china_str,
)
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from . import models, schemas
from .audit import log_audit
from .auth import get_current_user_optional, get_current_user
from .config import settings
from .database import get_db

router = APIRouter(prefix="/api/usage", tags=["usage"])

# 导出表头：含维护登记扩展列（登记日期、实际登记时间、设备、床号、开机/关机、维护人、设备状况等）
EXPORT_HEADERS = [
    "登记日期", "实际登记时间", "设备", "床号", "ID号", "姓名", "开机时间", "关机时间",
    "维护人", "使用类型", "设备状况", "日常保养", "终末消毒", "备注",
]

# 导出日期时间格式：中国时区展示（YYYY-MM-DD HH:mm:ss）
_DISPLAY_DATETIME_FMT = "%Y-%m-%d %H:%M:%S"


def _format_display_datetime(dt: Optional[datetime]) -> str:
    """格式化为中国时区展示用日期时间，用于导出与查询页一致。"""
    return utc_naive_to_china_str(dt, _DISPLAY_DATETIME_FMT)


def _get_usage_type_label_map(db: Session) -> dict:
    """使用类型编码 -> 中文显示名，用于导出与页面展示一致。"""
    out = {}
    for item in db.query(models.DictItem).filter(
        models.DictItem.dict_type == "usage_type",
        models.DictItem.is_deleted.is_(False),
    ).all():
        code = item.code if isinstance(item.code, str) else str(item.code)
        out[code] = item.label or code
    for k in (1, 2, 3, 4, 5):
        if str(k) not in out:
            out[str(k)] = str(k)
    return out


def _record_to_row(r: models.UsageRecord, usage_type_label_map: Optional[dict] = None) -> List[str]:
    """单行导出数据，含维护登记扩展列。"""
    device = r.device
    user = r.user
    device_display = ""
    if device:
        device_display = (device.name or "") + "（" + (device.device_code or "") + "）" if device.name else (device.device_code or "")
    usage_type_str = str(r.usage_type) if r.usage_type is not None else ""
    if usage_type_label_map:
        usage_type_str = usage_type_label_map.get(usage_type_str, usage_type_str)
    reg_date = getattr(r, "registration_date", None)
    reg_date_str = reg_date.strftime("%Y-%m-%d") if reg_date else ""
    end_time = getattr(r, "end_time", None)
    eq_cond = getattr(r, "equipment_condition", None) or ""
    if eq_cond == "normal":
        eq_cond = "正常"
    elif eq_cond == "abnormal":
        eq_cond = "异常"
    daily = getattr(r, "daily_maintenance", None) or ""
    if daily == "clean":
        daily = "清洁"
    elif daily == "disinfect":
        daily = "消毒"
    return [
        reg_date_str,
        _format_display_datetime(getattr(r, "created_at", None) or r.start_time),
        device_display,
        (getattr(r, "bed_number", None) or ""),
        (getattr(r, "id_number", None) or ""),
        (getattr(r, "patient_name", None) or ""),
        _format_display_datetime(r.start_time),
        _format_display_datetime(end_time),
        user.real_name if user else "",
        usage_type_str,
        eq_cond,
        daily,
        (getattr(r, "terminal_disinfection", None) or "").replace("\n", " "),
        (r.note or "").replace("\n", " "),
    ]


# 导出单次最大条数，避免百万级一次加载导致 OOM
EXPORT_MAX_RECORDS = 50_000


def _usage_query(
    db: Session,
    device_code: Optional[str] = None,
    dept: Optional[str] = None,
    user_id: Optional[int] = None,
    from_time: Optional[datetime] = None,
    to_time: Optional[datetime] = None,
    registration_date_from: Optional[date] = None,
    registration_date_to: Optional[date] = None,
    bed_number: Optional[str] = None,
):
    query = db.query(models.UsageRecord).filter(models.UsageRecord.is_deleted.is_(False))
    if device_code:
        query = query.filter(models.UsageRecord.device_code == device_code)
    if dept:
        query = query.join(models.UsageRecord.device).filter(models.Device.dept == dept)
    if user_id is not None:
        query = query.filter(models.UsageRecord.user_id == user_id)
    if from_time:
        query = query.filter(models.UsageRecord.start_time >= from_time)
    if to_time:
        query = query.filter(models.UsageRecord.start_time <= to_time)
    if registration_date_from is not None:
        query = query.filter(models.UsageRecord.registration_date >= registration_date_from)
    if registration_date_to is not None:
        query = query.filter(models.UsageRecord.registration_date <= registration_date_to)
    if bed_number:
        query = query.filter(models.UsageRecord.bed_number == bed_number.strip())
    return query.order_by(models.UsageRecord.start_time.desc())


def _fetch_export_records(
    db: Session,
    device_code: Optional[str] = None,
    dept: Optional[str] = None,
    user_id: Optional[int] = None,
    from_time: Optional[datetime] = None,
    to_time: Optional[datetime] = None,
    registration_date_from: Optional[date] = None,
    registration_date_to: Optional[date] = None,
    bed_number: Optional[str] = None,
    limit: Optional[int] = None,
    offset: int = 0,
):
    """获取导出用记录，可分批（limit/offset）。limit=None 表示不限制条数（调用方需保证不超过 EXPORT_MAX_RECORDS）。"""
    query = (
        _usage_query(
            db, device_code, dept, user_id, from_time, to_time,
            registration_date_from, registration_date_to, bed_number,
        )
        .options(
            joinedload(models.UsageRecord.device),
            joinedload(models.UsageRecord.user),
        )
    )
    if limit is not None:
        query = query.limit(limit).offset(offset)
    return query.all()

# 防重复：同一用户、同一设备在此秒数内的重复提交只保留一条
DUPLICATE_WINDOW_SECONDS = 10


@router.post(
    "",
    response_model=schemas.UsageRecordRead,
    status_code=status.HTTP_201_CREATED,
)
def create_usage_record(
    payload: schemas.UsageRecordCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    # 必须登录后登记，保证每条记录归属到对应用户，不同用户内容可区分
    user = current_user

    device = (
        db.query(models.Device)
        .filter(
            models.Device.device_code == payload.device_code,
            models.Device.is_active.is_(True),
        )
        .first()
    )
    if not device or getattr(device, "is_deleted", False):
        raise HTTPException(status_code=404, detail="设备不存在或已停用/已删除")

    # 防重复：短时间同一用户、同一设备只保留一条（不含已撤销），按中国当前时间计算
    cutoff = now_china_as_utc() - timedelta(seconds=DUPLICATE_WINDOW_SECONDS)
    existing = (
        db.query(models.UsageRecord)
        .filter(
            models.UsageRecord.user_id == user.id,
            models.UsageRecord.device_code == payload.device_code,
            models.UsageRecord.is_deleted.is_(False),
            models.UsageRecord.created_at >= cutoff,
        )
        .order_by(models.UsageRecord.created_at.desc())
        .first()
    )
    if existing:
        return existing

    data = payload.model_dump()
    data["usage_type"] = str(payload.usage_type)
    if not data.get("start_time"):
        data["start_time"] = now_china_as_utc()
    # 维护登记：若前端传了登记日期则用，否则用中国时区当天
    if data.get("registration_date") is None:
        data["registration_date"] = china_today()
    # 前端传入的 naive 时间视为中国时区，转为 UTC 存库
    for key in ("start_time", "end_time"):
        if data.get(key) is not None:
            data[key] = parse_naive_as_china_then_utc(data[key])
    photo_urls_list = data.pop("photo_urls", None)
    if photo_urls_list:
        data["photo_urls"] = ",".join(photo_urls_list)

    record = models.UsageRecord(
        user_id=user.id,
        **data,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.post("/{record_id}/undo", status_code=status.HTTP_204_NO_CONTENT)
def undo_usage_record(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """用户撤销自己的一条登记记录（软删除），仅本人可操作，且仅支持时间范围内的记录。"""
    record = db.get(models.UsageRecord, record_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="记录不存在")
    if record.user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="只能撤销自己的登记记录")
    if getattr(record, "is_deleted", False):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="该记录已撤销")
    window_hours = max(0, getattr(settings, "UNDO_WINDOW_HOURS", 24))
    cutoff = now_china_as_utc() - timedelta(hours=window_hours)
    if (record.created_at or record.start_time) < cutoff:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"仅支持撤销最近 {window_hours} 小时内的登记记录",
        )
    record.is_deleted = True
    db.commit()
    return None


def _list_usage_query(
    db: Session,
    current_user: models.User,
    device_code: Optional[str] = None,
    dept: Optional[str] = None,
    user_id: Optional[int] = None,
    from_time: Optional[datetime] = None,
    to_time: Optional[datetime] = None,
    registration_date_from: Optional[date] = None,
    registration_date_to: Optional[date] = None,
    bed_number: Optional[str] = None,
    include_deleted: bool = False,
):
    query = db.query(models.UsageRecord)
    if not include_deleted:
        query = query.filter(models.UsageRecord.is_deleted.is_(False))
    if current_user.role == "user" and user_id is None:
        query = query.filter(models.UsageRecord.user_id == current_user.id)
    elif user_id is not None:
        query = query.filter(models.UsageRecord.user_id == user_id)
    if device_code:
        query = query.filter(models.UsageRecord.device_code == device_code)
    if dept:
        query = query.join(models.UsageRecord.device).filter(models.Device.dept == dept)
    if from_time:
        query = query.filter(models.UsageRecord.start_time >= from_time)
    if to_time:
        query = query.filter(models.UsageRecord.start_time <= to_time)
    if registration_date_from is not None:
        query = query.filter(models.UsageRecord.registration_date >= registration_date_from)
    if registration_date_to is not None:
        query = query.filter(models.UsageRecord.registration_date <= registration_date_to)
    if bed_number:
        query = query.filter(models.UsageRecord.bed_number == bed_number.strip())
    return query.order_by(models.UsageRecord.start_time.desc())


@router.get("/count")
def count_usage_records(
    device_code: Optional[str] = Query(None),
    dept: Optional[str] = Query(None),
    user_id: Optional[int] = Query(None),
    from_time: Optional[datetime] = Query(None),
    to_time: Optional[datetime] = Query(None),
    registration_date_from: Optional[date] = Query(None, description="登记日期起"),
    registration_date_to: Optional[date] = Query(None, description="登记日期止"),
    bed_number: Optional[str] = Query(None, description="床号"),
    include_deleted: bool = Query(False, description="仅本人查看时有效，为 true 则含已撤销记录"),
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(get_current_user_optional),
):
    """返回符合条件的使用记录总数，用于分页与工作台统计。"""
    if current_user is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="请先登录后查看记录")
    from_time = parse_naive_as_china_then_utc(from_time) if from_time else None
    to_time = parse_naive_as_china_then_utc(to_time) if to_time else None
    allow_include_deleted = (user_id is None) and include_deleted
    query = _list_usage_query(
        db, current_user, device_code, dept, user_id, from_time, to_time,
        registration_date_from, registration_date_to, bed_number,
        include_deleted=allow_include_deleted,
    )
    return {"total": query.count()}


@router.get("", response_model=List[schemas.UsageRecordRead])
def list_usage_records(
    device_code: Optional[str] = Query(None, description="设备编号，与 devices.device_code 一致"),
    dept: Optional[str] = Query(None, description="设备科室"),
    user_id: Optional[int] = Query(None),
    from_time: Optional[datetime] = Query(None),
    to_time: Optional[datetime] = Query(None),
    registration_date_from: Optional[date] = Query(None, description="登记日期起"),
    registration_date_to: Optional[date] = Query(None, description="登记日期止"),
    bed_number: Optional[str] = Query(None, description="床号"),
    limit: int = Query(100, ge=1, le=500, description="每页条数"),
    offset: int = Query(0, ge=0, description="偏移量，用于分页"),
    include_deleted: bool = Query(False, description="仅本人查看时有效，为 true 则含已撤销记录"),
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(get_current_user_optional),
):
    if current_user is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="请先登录后查看记录",
        )
    from_time = parse_naive_as_china_then_utc(from_time) if from_time else None
    to_time = parse_naive_as_china_then_utc(to_time) if to_time else None
    allow_include_deleted = (user_id is None) and include_deleted
    query = (
        _list_usage_query(
            db, current_user, device_code, dept, user_id, from_time, to_time,
            registration_date_from, registration_date_to, bed_number,
            include_deleted=allow_include_deleted,
        )
        .options(
            joinedload(models.UsageRecord.device),
            joinedload(models.UsageRecord.user),
        )
    )
    records = query.offset(offset).limit(limit).all()
    return [
        schemas.UsageRecordRead.model_validate(r).model_copy(
            update={
                "device_name": r.device.name if r.device else None,
                "user_name": r.user.real_name if r.user else None,
                "is_deleted": getattr(r, "is_deleted", False),
            }
        )
        for r in records
    ]


def _build_excel(records: list, usage_type_label_map: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "使用记录"
    for col, h in enumerate(EXPORT_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, r in enumerate(records, 2):
        for col_idx, val in enumerate(_record_to_row(r, usage_type_label_map), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    # 表头加粗
    for c in range(1, len(EXPORT_HEADERS) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_pdf(records: list, usage_type_label_map: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    # 中文字体名，用于表格和标题
    font_name = "Helvetica"
    # 1) 优先尝试 ReportLab 内置 CID 字体（Adobe 亚洲语言包，若系统已安装）
    try:
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        font_name = "STSong-Light"
    except Exception:
        pass
    # 2) 若未注册成功，尝试系统 TTF/TTC 字体
    if font_name == "Helvetica":
        windir = os.environ.get("WINDIR", "C:\\Windows")
        fonts_dir = os.path.join(windir, "Fonts")
        # Windows 常见中文字体（.ttc 居多，.ttf 部分系统有）
        candidate_paths = [
            os.path.join(fonts_dir, "simsun.ttc"),
            os.path.join(fonts_dir, "msyh.ttc"),
            os.path.join(fonts_dir, "msyhbd.ttc"),
            os.path.join(fonts_dir, "simsun.ttf"),
            os.path.join(fonts_dir, "msyh.ttf"),
            os.path.join(fonts_dir, "simhei.ttf"),
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        ]
        # 项目内 static/fonts 下的字体（可放置 SimSun.ttf 等）
        static_fonts = os.path.join(os.path.dirname(__file__), "..", "static", "fonts")
        for name in ("SimSun.ttf", "SimHei.ttf", "msyh.ttf", "CJK.ttf"):
            candidate_paths.append(os.path.join(static_fonts, name))
        for path in candidate_paths:
            if not os.path.isfile(path):
                continue
            try:
                # TTC 需指定 subfontIndex（取第一个子字体）
                if path.lower().endswith(".ttc"):
                    pdfmetrics.registerFont(TTFont("CJK", path, subfontIndex=0))
                else:
                    pdfmetrics.registerFont(TTFont("CJK", path))
                font_name = "CJK"
                break
            except TypeError:
                # 旧版 reportlab 无 subfontIndex，对 .ttc 直接传 path 可能报错，跳过
                try:
                    pdfmetrics.registerFont(TTFont("CJK", path))
                    font_name = "CJK"
                    break
                except Exception:
                    pass
            except Exception:
                pass

    data = [EXPORT_HEADERS]
    for r in records:
        data.append(_record_to_row(r, usage_type_label_map))
    t = Table(data)
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0f2f1")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = font_name
    doc.build([Paragraph("使用记录导出", title_style), t])
    buf.seek(0)
    return buf.read()


def _export_csv_generator(
    db: Session,
    device_code, dept, user_id, from_time, to_time,
    registration_date_from: Optional[date] = None,
    registration_date_to: Optional[date] = None,
    bed_number: Optional[str] = None,
    usage_type_label_map: Optional[dict] = None,
):
    """流式生成 CSV：先表头，再分批查询写入。"""
    import codecs
    yield codecs.BOM_UTF8
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(EXPORT_HEADERS)
    yield output.getvalue()
    output.close()
    offset = 0
    batch_size = 5000
    while offset < EXPORT_MAX_RECORDS:
        batch = _fetch_export_records(
            db, device_code, dept, user_id, from_time, to_time,
            registration_date_from, registration_date_to, bed_number,
            limit=batch_size, offset=offset,
        )
        if not batch:
            break
        output = StringIO()
        writer = csv.writer(output)
        for r in batch:
            writer.writerow(_record_to_row(r, usage_type_label_map))
        yield output.getvalue()
        output.close()
        offset += len(batch)
        if len(batch) < batch_size:
            break


@router.get("/export")
def export_usage_records(
    device_code: Optional[str] = Query(None, description="设备编号"),
    dept: Optional[str] = Query(None, description="设备科室"),
    user_id: Optional[int] = Query(None),
    from_time: Optional[datetime] = Query(None),
    to_time: Optional[datetime] = Query(None),
    registration_date_from: Optional[date] = Query(None, description="登记日期起"),
    registration_date_to: Optional[date] = Query(None, description="登记日期止"),
    bed_number: Optional[str] = Query(None, description="床号"),
    format: str = Query("csv", description="导出格式: csv / xlsx / pdf"),
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(get_current_user_optional),
):
    if not current_user or current_user.role not in ("device_admin", "sys_admin"):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="仅管理员可导出",
        )
    from_time = parse_naive_as_china_then_utc(from_time) if from_time else None
    to_time = parse_naive_as_china_then_utc(to_time) if to_time else None
    total = _usage_query(
        db, device_code, dept, user_id, from_time, to_time,
        registration_date_from, registration_date_to, bed_number,
    ).count()
    if total > EXPORT_MAX_RECORDS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"符合条件记录共 {total} 条，超过单次导出上限 {EXPORT_MAX_RECORDS} 条，请缩小时间范围或筛选条件后导出。",
        )
    fmt = (format or "csv").lower().strip()
    usage_type_label_map = _get_usage_type_label_map(db)

    if fmt == "csv":
        log_audit(db, current_user.id, "usage.export", None, None, f"format={fmt},count={total}")
        return StreamingResponse(
            _export_csv_generator(
                db, device_code, dept, user_id, from_time, to_time,
                registration_date_from, registration_date_to, bed_number,
                usage_type_label_map,
            ),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="usage_records.csv"'},
        )
    records = _fetch_export_records(
        db, device_code, dept, user_id, from_time, to_time,
        registration_date_from, registration_date_to, bed_number,
        limit=EXPORT_MAX_RECORDS,
    )
    log_audit(db, current_user.id, "usage.export", None, None, f"format={fmt},count={len(records)}")
    if fmt == "xlsx":
        content = _build_excel(records, usage_type_label_map)
        filename = "usage_records.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif fmt == "pdf":
        content = _build_pdf(records, usage_type_label_map)
        filename = "usage_records.pdf"
        media_type = "application/pdf"
    else:
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(EXPORT_HEADERS)
        for r in records:
            writer.writerow(_record_to_row(r, usage_type_label_map))
        content = output.getvalue().encode("utf-8-sig")
        output.close()
        filename = "usage_records.csv"
        media_type = "text/csv; charset=utf-8"
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

