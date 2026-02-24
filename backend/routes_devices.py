import csv
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional

import qrcode
from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status
from sqlalchemy import inspect
from sqlalchemy.orm import Session

from . import models, schemas
from .audit import log_audit
from .auth import get_current_user_optional, require_role
from .config import settings
from .database import engine, get_db
from .device_code_utils import normalize_device_code

# 设备导出表头
DEVICE_EXPORT_HEADERS = [
    "设备编号", "设备名称", "科室", "位置", "状态", "启用", "已删除", "创建时间",
]

# 缓存：devices 表是否有 is_deleted 列（未迁移的旧库没有）
_devices_has_is_deleted: Optional[bool] = None


def _devices_table_has_is_deleted() -> bool:
    global _devices_has_is_deleted
    if _devices_has_is_deleted is None:
        try:
            cols = [c["name"] for c in inspect(engine).get_columns("devices")]
            _devices_has_is_deleted = "is_deleted" in cols
        except Exception:
            _devices_has_is_deleted = False
    return _devices_has_is_deleted


router = APIRouter(prefix="/api/devices", tags=["devices"])


@router.post(
    "",
    response_model=schemas.DeviceRead,
    status_code=status.HTTP_201_CREATED,
)
def create_device(
    payload: schemas.DeviceCreate,
    db: Session = Depends(get_db),
    current_user=Depends(require_role("device_admin", "sys_admin")),
):
    # 设备编号、名称、科室均为必填
    if not (payload.device_code and str(payload.device_code).strip()):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="设备编号为必填项",
        )
    if not (payload.name and str(payload.name).strip()):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="设备名称为必填项",
        )
    if not (payload.dept is not None and str(payload.dept).strip()):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="科室为必填项",
        )
    # 检查 device_code 唯一
    existed = (
        db.query(models.Device)
        .filter(models.Device.device_code == payload.device_code)
        .first()
    )
    if existed:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="设备编号已存在",
        )

    data = payload.model_dump()
    data["status"] = str(payload.status)
    device = models.Device(**data)
    db.add(device)
    db.flush()  # 获得 device.id，与审计同事务提交
    log_audit(
        db,
        current_user.id,
        "device.create",
        "device",
        device.id,
        f"device_code={device.device_code}",
        do_commit=False,
    )
    db.commit()
    db.refresh(device)
    return device


def _devices_query(
    db: Session,
    dept: Optional[str] = None,
    q: Optional[str] = None,
    include_inactive: bool = False,
    include_deleted: bool = False,
    deleted_only: bool = False,
    inactive_only: bool = False,
    is_admin: bool = False,
):
    query = db.query(models.Device)
    if _devices_table_has_is_deleted() and deleted_only and is_admin:
        query = query.filter(models.Device.is_deleted.is_(True))
        # “只显示已删除”时不按启用状态过滤，便于查看全部已删除
    elif inactive_only and is_admin:
        query = query.filter(models.Device.is_active.is_(False))
        if _devices_table_has_is_deleted():
            query = query.filter(models.Device.is_deleted.is_(False))
    else:
        if not (include_inactive and is_admin):
            query = query.filter(models.Device.is_active.is_(True))
        if _devices_table_has_is_deleted() and not (include_deleted and is_admin):
            query = query.filter(models.Device.is_deleted.is_(False))
    if dept:
        query = query.filter(models.Device.dept == dept)
    if q:
        like = f"%{q}%"
        query = query.filter(
            (models.Device.name.ilike(like))
            | (models.Device.device_code.ilike(like))
        )
    return query.order_by(models.Device.id.desc())


@router.get("/suggest")
def suggest_devices(
    q: Optional[str] = Query(None, description="名称或编号模糊搜索，为空则返回最近一批"),
    limit: int = Query(30, ge=1, le=100),
    dept: Optional[str] = Query(None, description="仅返回该科室设备，用于科室下拉联想"),
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(get_current_user_optional),
):
    """联想/下拉用：轻量返回设备列表，用于使用记录筛选等。数据量大时避免一次拉全量。"""
    q_normalized = normalize_device_code(q) if q else q
    query = db.query(models.Device).filter(models.Device.is_active.is_(True))
    if _devices_table_has_is_deleted():
        query = query.filter(models.Device.is_deleted.is_(False))
    if dept:
        query = query.filter(models.Device.dept == dept)
    if q_normalized and q_normalized.strip():
        like = f"%{q_normalized.strip()}%"
        query = query.filter(
            (models.Device.name.ilike(like))
            | (models.Device.device_code.ilike(like))
        )
    rows = query.order_by(models.Device.id.desc()).limit(limit).all()
    return [
        {"id": d.id, "device_code": d.device_code, "name": d.name, "dept": d.dept or ""}
        for d in rows
    ]


@router.get("/count")
def count_devices(
    dept: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    include_inactive: bool = Query(False),
    include_deleted: bool = Query(False),
    deleted_only: bool = Query(False, description="管理员可传 true 仅统计已删除设备"),
    inactive_only: bool = Query(False, description="管理员可传 true 仅统计已停用设备"),
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(get_current_user_optional),
):
    """返回符合条件的设备总数，用于分页展示。"""
    is_admin = current_user and current_user.role in ("device_admin", "sys_admin")
    query = _devices_query(db, dept, q, include_inactive, include_deleted, deleted_only, inactive_only, is_admin)
    return {"total": query.count()}


@router.get("", response_model=List[schemas.DeviceRead])
def list_devices(
    dept: Optional[str] = Query(None),
    q: Optional[str] = Query(
        None, description="按名称或编号模糊搜索"
    ),
    include_inactive: bool = Query(False, description="管理员可传 true 查看含已停用设备"),
    include_deleted: bool = Query(False, description="管理员可传 true 查看全部（含已删除），与 deleted_only 二选一"),
    deleted_only: bool = Query(False, description="管理员可传 true 仅查看已删除设备"),
    inactive_only: bool = Query(False, description="管理员可传 true 仅查看已停用设备"),
    limit: int = Query(100, ge=1, le=500, description="每页条数"),
    offset: int = Query(0, ge=0, description="偏移量，用于分页"),
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(get_current_user_optional),
):
    is_admin = current_user and current_user.role in ("device_admin", "sys_admin")
    q_normalized = normalize_device_code(q) if q else q
    query = _devices_query(db, dept, q_normalized, include_inactive, include_deleted, deleted_only, inactive_only, is_admin)
    return query.offset(offset).limit(limit).all()


# 设备状态默认中文（字典表为空或未匹配时兜底）
_DEVICE_STATUS_DEFAULT_LABELS = {
    "1": "可用",
    "2": "使用中",
    "3": "维修中",
    "4": "故障",
    "5": "报废",
}


def _device_to_export_row(d: models.Device, status_label_map: dict) -> List[str]:
    """单行设备导出数据，状态为中文。"""
    status_code = getattr(d, "status", None)
    if status_code is None:
        status_code = "1"
    status_code_str = str(status_code).strip()
    status_label = (
        status_label_map.get(status_code_str)
        or (status_label_map.get(int(status_code_str)) if status_code_str.isdigit() else None)
        or _DEVICE_STATUS_DEFAULT_LABELS.get(status_code_str)
        or status_code_str
    )
    return [
        d.device_code or "",
        d.name or "",
        d.dept or "",
        d.location or "",
        status_label,
        "是" if d.is_active else "否",
        "是" if getattr(d, "is_deleted", False) else "否",
        (d.created_at.strftime("%Y-%m-%d %H:%M:%S") if d.created_at else ""),
    ]


@router.get("/export")
def export_devices(
    dept: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    include_inactive: bool = Query(True, description="导出含已停用设备"),
    include_deleted: bool = Query(False, description="导出含已删除设备"),
    deleted_only: bool = Query(False, description="仅导出已删除设备"),
    inactive_only: bool = Query(False, description="仅导出已停用设备"),
    format: str = Query("csv", description="导出格式: csv / xlsx"),
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(get_current_user_optional),
):
    """导出设备列表（CSV/Excel），便于核对；仅管理员可导出。"""
    if not current_user or current_user.role not in ("device_admin", "sys_admin"):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="仅管理员可导出",
        )
    is_admin = True
    query = _devices_query(db, dept, q, include_inactive, include_deleted, deleted_only, inactive_only, is_admin)
    devices = query.all()
    status_label_map = {}
    for item in db.query(models.DictItem).filter(
        models.DictItem.dict_type == "device_status",
        models.DictItem.is_deleted.is_(False),
    ).all():
        label = (getattr(item, "label", None) or str(item.code) or "").strip()
        if not label:
            label = str(item.code) if item.code is not None else ""
        code_raw = getattr(item, "code", None)
        if code_raw is not None:
            code_key = str(code_raw).strip()
            status_label_map[code_key] = label
            if code_key.isdigit():
                status_label_map[int(code_key)] = label
    for k, v in _DEVICE_STATUS_DEFAULT_LABELS.items():
        if k not in status_label_map:
            status_label_map[k] = v
    rows = [DEVICE_EXPORT_HEADERS] + [_device_to_export_row(d, status_label_map) for d in devices]
    fmt = (format or "csv").lower().strip()
    log_audit(db, current_user.id, "device.export", None, None, f"format={fmt},count={len(devices)}")
    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "设备列表"
        for i, row in enumerate(rows, 1):
            for j, cell in enumerate(row, 1):
                ws.cell(row=i, column=j, value=cell)
        if rows:
            for cell in ws[1]:
                cell.font = Font(bold=True)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="devices.xlsx"'},
        )
    output = StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)
    content = output.getvalue().encode("utf-8-sig")
    output.close()
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="devices.csv"'},
    )


# 批量导入表头（与模板一致）
IMPORT_HEADERS = ["设备编号", "设备名称", "科室", "位置", "状态"]


@router.get("/import-template")
def download_import_template(
    _user=Depends(require_role("device_admin", "sys_admin")),
):
    """下载设备批量导入 Excel 模板（含表头与示例行）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "设备导入"
    for col, h in enumerate(IMPORT_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=1, value="DEV-001")
    ws.cell(row=2, column=2, value="示例设备")
    ws.cell(row=2, column=3, value="内科")
    ws.cell(row=2, column=4, value="一楼")
    ws.cell(row=2, column=5, value="1")
    for c in range(1, len(IMPORT_HEADERS) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="devices_import_template.xlsx"'},
    )


@router.post("/import", response_model=Dict[str, Any])
def import_devices(
    file: UploadFile = File(..., description="Excel 文件（.xlsx），表头：设备编号、设备名称、科室、位置、状态）"),
    db: Session = Depends(get_db),
    current_user=Depends(require_role("device_admin", "sys_admin")),
):
    """批量导入设备：Excel 首行为表头，从第二行起为数据；设备编号重复则跳过该行。"""
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的 Excel 文件")
    content = file.file.read()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法解析 Excel：{e!s}")
    ws = wb.active
    if not ws:
        raise HTTPException(status_code=400, detail="Excel 无有效工作表")
    created = 0
    skipped = 0
    errors: List[str] = []
    seen_codes: set[str] = set()
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            continue
        code = (row[0] if len(row) > 0 else None)
        name = (row[1] if len(row) > 1 else None)
        dept = (row[2] if len(row) > 2 else None)
        location = (row[3] if len(row) > 3 else None) or None
        status_val = (row[4] if len(row) > 4 else None) or "1"
        code = str(code).strip() if code is not None else ""
        name = str(name).strip() if name is not None else ""
        dept = str(dept).strip() if dept is not None else ""
        location = str(location).strip() if location else None
        if not code:
            errors.append(f"第{row_idx}行：设备编号为空，已跳过")
            skipped += 1
            continue
        if not name:
            errors.append(f"第{row_idx}行：设备名称为空，已跳过")
            skipped += 1
            continue
        if not dept:
            errors.append(f"第{row_idx}行：科室为空，已跳过")
            skipped += 1
            continue
        if code in seen_codes:
            errors.append(f"第{row_idx}行：设备编号 {code} 在本文件中重复，已跳过")
            skipped += 1
            continue
        seen_codes.add(code)
        status_str = str(status_val).strip() if status_val else "1"
        if status_str.isdigit():
            pass
        else:
            status_str = "1"
        existed = db.query(models.Device).filter(models.Device.device_code == code).first()
        if existed:
            errors.append(f"第{row_idx}行：设备编号 {code} 已存在，已跳过")
            skipped += 1
            continue
        device = models.Device(
            device_code=code,
            name=name,
            dept=dept,
            location=location,
            status=status_str,
            is_active=True,
            is_deleted=False,
        )
        db.add(device)
        db.flush()
        log_audit(db, current_user.id, "device.create", "device", device.id, f"device_code={code},import", do_commit=False)
        created += 1
    wb.close()
    db.commit()
    log_audit(db, current_user.id, "device.import", None, None, f"created={created},skipped={skipped}")
    return {"created": created, "skipped": skipped, "errors": errors}


@router.get("/{device_id}", response_model=schemas.DeviceRead)
def get_device(device_id: int, db: Session = Depends(get_db)):
    device = db.get(models.Device, device_id)
    if not device or getattr(device, "is_deleted", False) or not device.is_active:
        raise HTTPException(status_code=404, detail="设备不存在")
    return device


@router.get("/{device_id}/qrcode")
def get_device_qrcode(
    device_id: int,
    db: Session = Depends(get_db),
):
    """
    返回设备二维码 PNG 图片，后续可用于打印标签。
    二维码内容建议为短链接或包含设备唯一信息的 URL。
    """
    device = db.get(models.Device, device_id)
    if not device or getattr(device, "is_deleted", False) or not device.is_active:
        raise HTTPException(status_code=404, detail="设备不存在")

    # 二维码内容必须为完整 URL，否则企微/浏览器扫码无法打开
    path = f"/h5/scan?device_code={device.device_code}"
    if device.qr_value and (device.qr_value.startswith("http://") or device.qr_value.startswith("https://")):
        qr_value = device.qr_value
    else:
        relative = (device.qr_value or path).strip()
        if not relative.startswith("/"):
            relative = "/" + relative
        qr_value = f"{settings.BASE_URL.rstrip('/')}{relative}"

    img = qrcode.make(qr_value)
    buf = BytesIO()
    # 兼容 PIL 与 pypng 等后端：PIL 用 format="PNG"，pypng 只支持 .save(buf)
    try:
        img.save(buf, format="PNG")
    except TypeError:
        img.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="image/png",
        headers={
            "Content-Disposition": f'inline; filename="device_{device.id}_qrcode.png"'
        },
    )


@router.patch("/{device_id}", response_model=schemas.DeviceRead)
def update_device(
    device_id: int,
    payload: schemas.DeviceUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(require_role("device_admin", "sys_admin")),
):
    """更新设备（编号、名称、科室、位置、状态、启用、软删除）。仅管理员。"""
    device = db.get(models.Device, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    was_deleted = getattr(device, "is_deleted", False) if _devices_table_has_is_deleted() else False
    if payload.device_code is not None:
        new_code = (payload.device_code or "").strip()
        if not new_code:
            raise HTTPException(status_code=400, detail="设备编号不能为空")
        if new_code != (device.device_code or ""):
            existed = db.query(models.Device).filter(models.Device.device_code == new_code).first()
            if existed and existed.id != device_id:
                raise HTTPException(status_code=400, detail="设备编号已存在")
            device.device_code = new_code
    if payload.name is not None:
        device.name = payload.name.strip()
    if payload.dept is not None:
        device.dept = payload.dept.strip() or None
    if payload.location is not None:
        device.location = payload.location.strip() or None
    if payload.status is not None:
        device.status = str(payload.status)
    if payload.is_active is not None:
        device.is_active = payload.is_active
    if payload.is_deleted is not None and _devices_table_has_is_deleted():
        device.is_deleted = payload.is_deleted
    code_prefix = f"device_code={device.device_code or ''}"
    if payload.is_deleted is True:
        log_audit(
            db,
            current_user.id,
            "device.delete",
            "device",
            device_id,
            f"{code_prefix},soft_delete",
            do_commit=False,
        )
    elif payload.is_deleted is False and was_deleted:
        log_audit(db, current_user.id, "device.restore", "device", device_id, code_prefix, do_commit=False)
    else:
        parts = [k for k in ("name", "dept", "status", "is_active") if getattr(payload, k) is not None]
        details = code_prefix + ("," + ",".join(parts) if parts else "")
        log_audit(db, current_user.id, "device.update", "device", device_id, details, do_commit=False)
    db.commit()
    db.refresh(device)
    return device

